#!/usr/bin/env python3
# /// script
# dependencies = ["openpyxl>=3.1.2"]
# ///
"""Build an Obsidian-friendly Challenge 2 knowledge base.

The script leaves raw source folders untouched and regenerates the LLM-wiki layer
under challenge-2/wiki/.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import html.parser
import json
import os
import re
import shutil
import subprocess
import sys
import textwrap
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote
import xml.etree.ElementTree as ET

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised when not run through uv.
    openpyxl = None


ROOT = Path(__file__).resolve().parents[1]
STRUCTURED_DIR = ROOT / "structured_files"
UNSTRUCTURED_DIR = ROOT / "unstructured_files"
WIKI_DIR = ROOT / "wiki"
DATA_DIR = WIKI_DIR / "data"
SOURCES_DIR = WIKI_DIR / "sources"
TOPICS_DIR = WIKI_DIR / "topics"
ENTITIES_DIR = WIKI_DIR / "entities"
MAPS_DIR = WIKI_DIR / "maps"
TABLES_DIR = DATA_DIR / "tables"

TODAY = dt.date.today().isoformat()
BUILD_TS = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()


TOPIC_DEFS = [
    {
        "slug": "housing-benefit",
        "title": "Housing Benefit",
        "aliases": ["Housing Benefit", "HB"],
        "keywords": ["housing benefit", "hb1", "local authority housing", "eligible rent"],
    },
    {
        "slug": "discretionary-housing-payments",
        "title": "Discretionary Housing Payments",
        "aliases": ["DHP", "DHPs", "Discretionary Housing Payments"],
        "keywords": ["discretionary housing payment", "dhp", "shortfall"],
    },
    {
        "slug": "council-tax-reduction",
        "title": "Council Tax Reduction",
        "aliases": ["CTR", "Council Tax Reduction"],
        "keywords": ["council tax reduction", "ctr", "billing authority"],
    },
    {
        "slug": "homelessness-and-social-housing",
        "title": "Homelessness And Social Housing",
        "aliases": ["Homelessness", "Social Housing", "Right to Buy"],
        "keywords": ["homelessness", "social housing", "right to buy", "duty to refer"],
    },
    {
        "slug": "small-business-and-self-employment",
        "title": "Small Business And Self-employment",
        "aliases": ["Small Business", "Self-employment"],
        "keywords": ["self-employed", "self employment", "small business", "starting a business", "registering as self-employed"],
    },
    {
        "slug": "employment-rights-and-flexible-working",
        "title": "Employment Rights And Flexible Working",
        "aliases": ["Employment Rights", "Flexible Working", "SSP"],
        "keywords": ["employment rights", "flexible working", "statutory sick pay", "ssp", "minimum wage"],
    },
    {
        "slug": "procurement-and-spending-controls",
        "title": "Procurement And Spending Controls",
        "aliases": ["Procurement", "Spending Controls", "Approval Matrix"],
        "keywords": ["procurement", "spending controls", "approval matrix", "cabinet office spending controls", "business case"],
    },
    {
        "slug": "data-protection-and-information-security",
        "title": "Data Protection And Information Security",
        "aliases": ["Data Protection", "Information Security", "Acceptable Use"],
        "keywords": ["data protection", "information security", "acceptable use", "personal data", "uk gdpr"],
    },
    {
        "slug": "foi-and-transparency",
        "title": "FOI And Transparency",
        "aliases": ["FOI", "Freedom of Information", "Transparency"],
        "keywords": ["foi", "freedom of information", "response template", "disclosure"],
    },
    {
        "slug": "people-policies",
        "title": "People Policies",
        "aliases": ["HR Policies", "People Services"],
        "keywords": ["annual leave", "grievance", "recruitment", "performance", "whistleblowing", "travel", "subsistence"],
    },
    {
        "slug": "incident-risk-and-assurance",
        "title": "Incident Risk And Assurance",
        "aliases": ["Incident Reporting", "Risk", "Assurance", "Compliance"],
        "keywords": ["incident", "risk", "assurance", "compliance", "equality impact", "programme board"],
    },
    {
        "slug": "overpayment-recovery",
        "title": "Overpayment Recovery",
        "aliases": ["Overpayment", "Recovery Procedures"],
        "keywords": ["overpayment", "recovery", "fraud", "non-fraud"],
    },
    {
        "slug": "welsh-language-standards",
        "title": "Welsh Language Standards",
        "aliases": ["Welsh Language", "Welsh Language Standards"],
        "keywords": ["welsh language", "welsh", "standards"],
    },
    {
        "slug": "staff-directory",
        "title": "Staff Directory",
        "aliases": ["Staff Directory", "Directory Extract"],
        "keywords": ["staff directory", "directory extract"],
    },
]

ENTITY_DEFS = [
    {"slug": "dluhc", "title": "Department For Levelling Up Housing And Communities", "aliases": ["DLUHC"], "keywords": ["dluhc", "levelling up"]},
    {"slug": "dwp", "title": "Department For Work And Pensions", "aliases": ["DWP"], "keywords": ["dwp", "work and pensions"]},
    {"slug": "dbt", "title": "Department For Business And Trade", "aliases": ["DBT"], "keywords": ["department for business and trade", "dbt"]},
    {"slug": "hmrc", "title": "HM Revenue And Customs", "aliases": ["HMRC"], "keywords": ["hm revenue", "hmrc"]},
    {"slug": "local-authorities", "title": "Local Authorities", "aliases": ["Local authority", "Local authorities"], "keywords": ["local authority", "local authorities", "billing authority"]},
    {"slug": "housing-act-1996", "title": "Housing Act 1996", "aliases": ["Housing Act 1996"], "keywords": ["housing act 1996"]},
    {"slug": "homelessness-reduction-act-2017", "title": "Homelessness Reduction Act 2017", "aliases": ["HRA 2017"], "keywords": ["homelessness reduction act 2017"]},
    {"slug": "employment-rights-act-1996", "title": "Employment Rights Act 1996", "aliases": ["ERA 1996"], "keywords": ["employment rights act 1996", "era 1996"]},
    {"slug": "hb1-form", "title": "HB1 Form", "aliases": ["HB1"], "keywords": ["hb1", "claim form"]},
    {"slug": "rtb1-form", "title": "RTB1 Form", "aliases": ["RTB1"], "keywords": ["rtb1"]},
    {"slug": "universal-credit-migration", "title": "Universal Credit Migration", "aliases": ["UC Migration"], "keywords": ["universal credit migration", "uc migration"]},
    {"slug": "cabinet-office", "title": "Cabinet Office", "aliases": ["Cabinet Office"], "keywords": ["cabinet office"]},
    {"slug": "finance-directorate", "title": "Finance Directorate", "aliases": ["Finance Directorate"], "keywords": ["finance directorate", "finance business partner"]},
    {"slug": "people-services", "title": "People Services", "aliases": ["People Services"], "keywords": ["people services"]},
]

MAP_DEFS = [
    {
        "slug": "housing-and-benefits",
        "title": "Housing And Benefits Map",
        "topics": ["housing-benefit", "discretionary-housing-payments", "council-tax-reduction", "homelessness-and-social-housing", "overpayment-recovery"],
    },
    {
        "slug": "small-business-and-employment",
        "title": "Small Business And Employment Map",
        "topics": ["small-business-and-self-employment", "employment-rights-and-flexible-working"],
    },
    {
        "slug": "corporate-operations",
        "title": "Corporate Operations Map",
        "topics": ["procurement-and-spending-controls", "data-protection-and-information-security", "foi-and-transparency", "staff-directory"],
    },
    {
        "slug": "people-policies",
        "title": "People Policies Map",
        "topics": ["people-policies", "employment-rights-and-flexible-working", "staff-directory"],
    },
    {
        "slug": "risk-assurance-and-compliance",
        "title": "Risk Assurance And Compliance Map",
        "topics": ["incident-risk-and-assurance", "data-protection-and-information-security", "welsh-language-standards", "foi-and-transparency"],
    },
]

DOC_ID_RE = re.compile(r"\bDOC-[A-Z]{2}-\d{3}[A-Za-z]?\b")
MARKDOWN_LINK_RE = re.compile(r"\[([^\]\n]+)\]\(([^)\n]+)\)")
SYNTHETIC_DATA_NOTICE = "Challenge 2 corpus data is synthetic hackathon fixture data."
GREEN_BOOK_URL = "https://www.gov.uk/government/publications/the-green-book-appraisal-and-evaluation-in-central-government"


@dataclass
class SourceRecord:
    source_id: str
    title: str
    source_path: Path
    source_format: str
    note_path: Path
    document_type: str | None = None
    department: str | None = None
    owner: str | None = None
    status: str | None = None
    version: str | None = None
    publication_date: str | None = None
    last_updated: str | None = None
    audience: list[str] = field(default_factory=list)
    topics: list[str] = field(default_factory=list)
    supersedes: list[str] = field(default_factory=list)
    related_sources: list[str] = field(default_factory=list)
    raw_metadata: dict[str, Any] = field(default_factory=dict)
    technical_metadata: dict[str, Any] = field(default_factory=dict)
    extraction_method: str = "unknown"
    extraction_quality: str = "medium"
    extraction_warnings: list[str] = field(default_factory=list)
    contains_synthetic_identifiers: bool = False
    sensitivity_classification: str | None = None
    extracted_markdown: str = ""
    extracted_tables: list[dict[str, Any]] = field(default_factory=list)
    extracted_links: list[dict[str, str]] = field(default_factory=list)
    worksheets: list[dict[str, Any]] = field(default_factory=list)
    sha256: str = ""
    size_bytes: int = 0
    flags: list[str] = field(default_factory=list)
    matched_topics: list[str] = field(default_factory=list)
    matched_entities: list[str] = field(default_factory=list)


class MetaParser(html.parser.HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_title = False
        self.title_parts: list[str] = []
        self.meta: dict[str, str] = {}

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = {k.lower(): v or "" for k, v in attrs}
        if tag.lower() == "title":
            self.in_title = True
        if tag.lower() == "meta":
            key = attrs_dict.get("name") or attrs_dict.get("property")
            content = attrs_dict.get("content")
            if key and content is not None:
                self.meta[key.lower()] = content.strip()

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "title":
            self.in_title = False

    def handle_data(self, data: str) -> None:
        if self.in_title:
            self.title_parts.append(data)

    @property
    def title(self) -> str | None:
        title = " ".join(p.strip() for p in self.title_parts if p.strip())
        return title or None


def run_cmd(args: list[str], input_text: str | None = None) -> str:
    try:
        result = subprocess.run(args, input=input_text, text=True, capture_output=True, check=False)
    except FileNotFoundError:
        return ""
    if result.returncode != 0:
        return result.stdout.strip() or result.stderr.strip()
    return result.stdout


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def slugify(value: str, max_len: int = 80) -> str:
    value = value.lower()
    value = re.sub(r"&", " and ", value)
    value = re.sub(r"[^a-z0-9]+", "-", value).strip("-")
    value = re.sub(r"-+", "-", value)
    return (value[:max_len].strip("-") or "untitled")


def unique(items: list[str | None]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for item in items:
        if item is None:
            continue
        value = str(item).strip()
        if not value:
            continue
        key = value.lower()
        if key not in seen:
            seen.add(key)
            out.append(value)
    return out


def rel_link(from_path: Path, to_path: Path, label: str | None = None) -> str:
    rel = os.path.relpath(to_path, from_path.parent)
    rel = "/".join(quote(part) for part in Path(rel).parts)
    return f"[{label or to_path.stem}]({rel})"


def yaml_scalar(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value)
    if text == "":
        return '""'
    return json.dumps(text, ensure_ascii=False)


def yaml_dump(data: dict[str, Any], indent: int = 0) -> str:
    lines: list[str] = []
    pad = " " * indent
    for key, value in data.items():
        if isinstance(value, dict):
            lines.append(f"{pad}{key}:")
            lines.append(yaml_dump(value, indent + 2))
        elif isinstance(value, list):
            if not value:
                lines.append(f"{pad}{key}: []")
            else:
                lines.append(f"{pad}{key}:")
                for item in value:
                    if isinstance(item, dict):
                        lines.append(f"{pad}  -")
                        lines.append(yaml_dump(item, indent + 4))
                    else:
                        lines.append(f"{pad}  - {yaml_scalar(item)}")
        else:
            lines.append(f"{pad}{key}: {yaml_scalar(value)}")
    return "\n".join(lines)


def frontmatter(data: dict[str, Any]) -> str:
    return "---\n" + yaml_dump(data) + "\n---\n\n"


def parse_simple_frontmatter(text: str) -> tuple[dict[str, Any], str]:
    if not text.startswith("---"):
        return {}, text
    parts = text.split("---", 2)
    if len(parts) < 3:
        return {}, text
    raw = parts[1]
    body = parts[2].lstrip()
    meta: dict[str, Any] = {}
    current_key: str | None = None
    for line in raw.splitlines():
        if not line.strip():
            continue
        if re.match(r"^[A-Za-z0-9_-]+:", line):
            key, value = line.split(":", 1)
            key = key.strip()
            value = value.strip()
            current_key = key
            if value == "":
                meta[key] = []
            else:
                meta[key] = value.strip('"').strip("'")
        elif current_key and line.strip().startswith("-"):
            if not isinstance(meta.get(current_key), list):
                meta[current_key] = []
            meta[current_key].append(line.strip()[1:].strip().strip('"').strip("'"))
    return meta, body


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def first_meaningful_line(text: str, fallback: str) -> str:
    skip = {
        "official",
        "draft",
        "confidential",
        "department for work and pensions",
        "department for business and trade",
        "department for levelling up, housing and communities",
        "hm revenue and customs",
    }
    for line in text.splitlines():
        cleaned = re.sub(r"^[#*\s>.-]+|[*\s]+$", "", line).strip()
        cleaned = re.sub(r"^\\?\d+\\?\.\s*", "", cleaned).strip()
        if not cleaned:
            continue
        if cleaned.lower() in skip or cleaned.lower().startswith("official"):
            continue
        if re.fullmatch(r"[-=_]{3,}", cleaned):
            continue
        return cleaned[:160]
    return fallback


def meaningful_metadata_value(*values: Any) -> str | None:
    placeholders = {"", "(anonymous)", "(unspecified)", "anonymous", "unspecified", "none", "null"}
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text.lower() in placeholders:
            continue
        return text
    return None


def extract_key_values(text: str) -> dict[str, str]:
    meta: dict[str, str] = {}
    patterns = {
        "document_id": r"^\s*(?:Document ID|Document Reference)\s*:\s*([A-Z0-9-]+)",
        "published": r"^\s*(?:Published|Date of Publication|Publication Date|Date)\s*:\s*([^\n]+)",
        "last_updated": r"^\s*(?:Last Updated|Updated)\s*:\s*([^\n]+)",
        "department": r"^\s*Department\s*:\s*([^\n]+)",
        "status": r"^\s*Status\s*:\s*([^\n]+)",
        "audience": r"^\s*Audience\s*:\s*([^\n]+)",
        "version": r"^\s*Version\s*[: ]\s*([^\n]+)",
        "owner": r"^\s*Owner\s*:\s*([^\n]+)",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            meta[key] = match.group(1).strip().strip(".")
    return meta


def normalise_status(value: str | None, path: Path, text: str) -> str | None:
    explicit = (value or "").strip().lower()
    head = text[:600].lower()
    if "draft" in explicit or "draft" in path.name.lower() or re.search(r"(?im)^\s*draft\b", text[:1000]):
        return "draft"
    if "superseded" in explicit:
        return "superseded"
    if "replaces the version" in head or "replaces the earlier" in head:
        return "current"
    if value:
        cleaned = value.strip().lower()
        if "current" in cleaned:
            return "current"
        return cleaned
    return "unknown"


def infer_document_type(path: Path, meta: dict[str, Any], text: str) -> str:
    explicit = meta.get("type") or meta.get("document-type") or meta.get("document_type")
    if explicit:
        return str(explicit)
    if path.suffix.lower() == ".xlsx":
        return "spreadsheet"
    name = path.stem.lower()
    blob = f"{name} {text[:1000].lower()}"
    if "template" in blob:
        return "template"
    if "minutes" in blob:
        return "minutes"
    if "policy" in blob:
        return "policy"
    if "guidance" in blob or "guide" in blob:
        return "guidance"
    if "framework" in blob:
        return "framework"
    if "assessment" in blob:
        return "assessment"
    if "survey" in blob or "statistics" in blob or "report" in blob:
        return "report"
    return "document"


def infer_department(meta: dict[str, Any], text: str) -> str | None:
    department = meta.get("department") or meta.get("author")
    if department:
        return str(department).strip()
    checks = [
        ("DLUHC", ["department for levelling up", "dluhc"]),
        ("DWP", ["department for work and pensions", "dwp"]),
        ("Department for Business and Trade", ["department for business and trade", "dbt"]),
        ("HM Revenue and Customs", ["hm revenue", "hmrc"]),
        ("Cabinet Office", ["cabinet office"]),
    ]
    lower = text[:5000].lower()
    for value, needles in checks:
        if any(n in lower for n in needles):
            return value
    return None


def infer_topics(text: str, meta: dict[str, Any]) -> list[str]:
    raw = meta.get("topics") or meta.get("topic")
    topics: list[str] = []
    if isinstance(raw, list):
        topics.extend(str(x) for x in raw)
    elif isinstance(raw, str):
        topics.extend(part.strip() for part in re.split(r"[,;]", raw))
    lower = text.lower()
    for topic in TOPIC_DEFS:
        if any(k in lower for k in topic["keywords"]):
            topics.append(topic["slug"])
    return unique([slugify(t, 50) for t in topics])


def find_related_sources(text: str, own_id: str | None = None) -> list[str]:
    ids = DOC_ID_RE.findall(text)
    return [i for i in unique(ids) if i != own_id]


def find_supersedes(meta: dict[str, Any], text: str) -> list[str]:
    values: list[str] = []
    raw = meta.get("supersedes") or meta.get("replaces")
    if isinstance(raw, list):
        values.extend(str(x) for x in raw)
    elif raw:
        values.append(str(raw))
    for pattern in [
        r"replaces the version published in ([^.]+)",
        r"supersedes[: ]+([^\n.]+)",
        r"replaces ([^.]+)",
    ]:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            values.append(match.group(0).strip())
    return unique(values)


def source_id_from_path(path: Path, meta: dict[str, Any], text: str) -> str:
    for key in ["document_id", "document-id", "document-id", "document reference", "document_reference"]:
        value = meta.get(key)
        if value:
            return str(value).strip()
    filename_match = DOC_ID_RE.search(path.stem.upper())
    if filename_match:
        return filename_match.group(0).upper()
    match = DOC_ID_RE.search(text[:3000].upper())
    if match:
        return match.group(0)
    return "UF-" + slugify(path.stem, 70).upper().replace("-", "-")


def pandoc(path: Path, source_format: str | None = None) -> str:
    args = ["pandoc", str(path), "-t", "gfm", "--wrap=none"]
    if source_format:
        args.insert(2, "-f")
        args.insert(3, source_format)
    return run_cmd(args)


def exif_metadata(path: Path) -> dict[str, Any]:
    output = run_cmd(["exiftool", "-json", str(path)])
    try:
        data = json.loads(output)
    except json.JSONDecodeError:
        return {}
    if not data:
        return {}
    result = dict(data[0])
    for key in ["SourceFile", "Directory", "FileAccessDate", "FileInodeChangeDate"]:
        result.pop(key, None)
    return result


def pdf_info(path: Path) -> dict[str, str]:
    output = run_cmd(["pdfinfo", str(path)])
    info: dict[str, str] = {}
    for line in output.splitlines():
        if ":" in line:
            key, value = line.split(":", 1)
            info[key.strip()] = value.strip()
    return info


def docx_core_props(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    try:
        with zipfile.ZipFile(path) as z:
            root = ET.fromstring(z.read("docProps/core.xml"))
    except Exception:
        return out
    for child in root:
        key = child.tag.split("}", 1)[-1]
        value = (child.text or "").strip()
        if value:
            out[key] = value
    return out


def markdown_table(rows: list[list[Any]], max_rows: int | None = None) -> str:
    cleaned: list[list[str]] = []
    width = 0
    for row in rows:
        values = ["" if value is None else str(value).replace("\n", " ").strip() for value in row]
        if any(values):
            cleaned.append(values)
            width = max(width, len(values))
    if not cleaned:
        return "_No populated rows found._\n"
    if max_rows:
        cleaned = cleaned[:max_rows]
    padded = [row + [""] * (width - len(row)) for row in cleaned]
    header = padded[0]
    body = padded[1:]
    def esc(value: str) -> str:
        return value.replace("|", "\\|")
    lines = ["| " + " | ".join(esc(v) for v in header) + " |"]
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in body:
        lines.append("| " + " | ".join(esc(v) for v in row) + " |")
    if max_rows and len(rows) > max_rows:
        lines.append(f"\n_Only the first {max_rows} populated rows are shown here._")
    return "\n".join(lines) + "\n"


def detect_fixed_width_tables(text: str) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    current: list[list[str]] = []
    blank_gap = 0

    def flush() -> None:
        nonlocal current, blank_gap
        if len(current) >= 2:
            width = max(len(row) for row in current)
            if width >= 2:
                tables.append({"rows": [row + [""] * (width - len(row)) for row in current]})
        current = []
        blank_gap = 0

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            if current:
                blank_gap += 1
                if blank_gap <= 1:
                    continue
            flush()
            continue
        cells = [cell.strip() for cell in re.split(r"\s{2,}", line) if cell.strip()]
        if len(cells) >= 2 and not line.startswith("•"):
            current.append(cells)
            blank_gap = 0
        else:
            flush()
    flush()
    return tables


def detect_extracted_links(text: str) -> list[dict[str, str]]:
    links: list[dict[str, str]] = []
    lower = text.lower()
    if "the-green-book-appraisal-and-evaluation" in lower and ("governent" in lower or "green book" in lower):
        links.append({
            "label": "HMT Green Book",
            "url": GREEN_BOOK_URL,
            "note": "Corrected from extracted PDF line break/typo; raw extracted text is preserved below.",
        })
    return links


def extract_xlsx(path: Path) -> tuple[str, dict[str, Any], list[dict[str, Any]], list[str]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required. Run with: uv run challenge-2/tools/build_wiki.py")
    wb = openpyxl.load_workbook(path, data_only=True)
    props = wb.properties
    meta = {
        "creator": props.creator,
        "created": props.created.isoformat() if props.created else None,
        "modified": props.modified.isoformat() if props.modified else None,
        "title": props.title,
        "subject": props.subject,
    }
    warnings: list[str] = []
    worksheets: list[dict[str, Any]] = []
    sections: list[str] = []
    for ws in wb.worksheets:
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        populated = [row for row in rows if any(v is not None and str(v).strip() for v in row)]
        preview = populated[:80]
        worksheets.append({
            "name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "populated_rows": len(populated),
            "rows": populated,
        })
        sections.append(f"## Worksheet: {ws.title}\n\n" + markdown_table(preview))
        if len(populated) > len(preview):
            warnings.append(f"Worksheet {ws.title} has {len(populated)} populated rows; note preview shows {len(preview)}.")
    return "\n".join(sections).strip() + "\n", meta, worksheets, warnings


def extract_source(path: Path) -> SourceRecord:
    suffix = path.suffix.lower().lstrip(".")
    raw_metadata: dict[str, Any] = {}
    text = ""
    method = suffix
    quality = "medium"
    warnings: list[str] = []
    worksheets: list[dict[str, Any]] = []

    if suffix == "md":
        raw = read_text(path)
        fm, body = parse_simple_frontmatter(raw)
        raw_metadata.update(fm)
        text = body
        method = "markdown-frontmatter"
        quality = "high"
    elif suffix == "html":
        raw = read_text(path)
        parser = MetaParser()
        parser.feed(raw)
        raw_metadata.update(parser.meta)
        if parser.title:
            raw_metadata["title"] = parser.title
        converted = pandoc(path, "html")
        text = converted or re.sub(r"<[^>]+>", " ", raw)
        method = "pandoc-html"
        quality = "high" if converted else "low"
        if not converted:
            warnings.append("pandoc did not return HTML conversion output; used stripped HTML fallback.")
    elif suffix == "txt":
        text = read_text(path)
        raw_metadata.update(extract_key_values(text[:3000]))
        method = "plain-text-regex"
        quality = "medium"
    elif suffix == "pdf":
        text = run_cmd(["pdftotext", "-layout", str(path), "-"])
        raw_metadata.update(pdf_info(path))
        method = "pdftotext-layout"
        quality = "medium" if text.strip() else "low"
        if not text.strip():
            warnings.append("No text extracted from PDF.")
    elif suffix == "docx":
        raw_metadata.update(docx_core_props(path))
        text = pandoc(path)
        method = "pandoc-docx"
        quality = "high" if text.strip() else "low"
        if not text.strip():
            warnings.append("pandoc did not return DOCX conversion output.")
    elif suffix == "xlsx":
        text, xlsx_meta, worksheets, xlsx_warnings = extract_xlsx(path)
        raw_metadata.update({k: v for k, v in xlsx_meta.items() if v})
        warnings.extend(xlsx_warnings)
        method = "openpyxl-workbook"
        quality = "high"
    else:
        text = read_text(path)
        warnings.append(f"Unhandled extension .{suffix}; read as text.")

    raw_metadata.update({f"exif_{k}": v for k, v in exif_metadata(path).items()})
    key_values = extract_key_values(text[:5000])
    for key, value in key_values.items():
        raw_metadata.setdefault(key, value)

    fallback_title = path.stem.replace("_", " ")
    metadata_title = meaningful_metadata_value(raw_metadata.get("title"), raw_metadata.get("Title"), raw_metadata.get("dc:title"))
    if suffix == "xlsx":
        title = metadata_title or fallback_title
    else:
        title = metadata_title or first_meaningful_line(text, fallback_title)
    source_id = source_id_from_path(path, raw_metadata, text)
    status = normalise_status(str(raw_metadata.get("status")) if raw_metadata.get("status") else None, path, text)
    department = infer_department(raw_metadata, text)
    document_type = infer_document_type(path, raw_metadata, text)
    publication_date = str(raw_metadata.get("publication_date") or raw_metadata.get("publication-date") or raw_metadata.get("published") or raw_metadata.get("date-published") or "").strip() or None
    last_updated = str(raw_metadata.get("last_updated") or raw_metadata.get("last-updated") or raw_metadata.get("last_updated") or "").strip() or None
    audience_raw = raw_metadata.get("audience")
    audience = []
    if isinstance(audience_raw, list):
        audience = [str(x) for x in audience_raw]
    elif audience_raw:
        audience = [str(audience_raw)]
    version = str(raw_metadata.get("version") or raw_metadata.get("Version") or "").strip() or None
    owner = str(raw_metadata.get("owner") or "").strip() or None
    topics = infer_topics(text + " " + title, raw_metadata)
    related_sources = find_related_sources(text, source_id)
    supersedes = find_supersedes(raw_metadata, text)
    contains_synthetic_identifiers = "staff directory" in path.stem.lower() or bool(
        re.search(r"\b(email|telephone|phone|line manager)\b", text[:2000], re.I)
        and "directory" in text[:2000].lower()
    )
    if re.search(r"OFFICIAL\s*[-—]\s*SENSITIVE", text[:2000], re.I):
        classification = "OFFICIAL-SENSITIVE"
    elif re.search(r"\bOFFICIAL\b", text[:2000], re.I):
        classification = "OFFICIAL"
    else:
        classification = None

    note_name = f"{slugify(source_id, 35)}-{slugify(title, 55)}.md"
    record = SourceRecord(
        source_id=source_id,
        title=title,
        source_path=path,
        source_format=suffix,
        note_path=SOURCES_DIR / note_name,
        document_type=document_type,
        department=department,
        owner=owner,
        status=status,
        version=version,
        publication_date=publication_date,
        last_updated=last_updated,
        audience=audience,
        topics=topics,
        supersedes=supersedes,
        related_sources=related_sources,
        raw_metadata=raw_metadata,
        technical_metadata={"size_bytes": path.stat().st_size, "sha256": sha256(path)},
        extraction_method=method,
        extraction_quality=quality,
        extraction_warnings=warnings,
        contains_synthetic_identifiers=contains_synthetic_identifiers,
        sensitivity_classification=classification,
        extracted_markdown=text.strip() + "\n",
        extracted_tables=detect_fixed_width_tables(text) if suffix == "pdf" else [],
        extracted_links=detect_extracted_links(text),
        worksheets=worksheets,
        sha256=sha256(path),
        size_bytes=path.stat().st_size,
    )
    record.flags = infer_flags(record)
    return record


def infer_flags(record: SourceRecord) -> list[str]:
    flags: list[str] = []
    blob = f"{record.source_path.name} {record.title} {record.extracted_markdown[:5000]}".lower()
    if record.status == "draft":
        flags.append("draft")
    if record.status == "superseded":
        flags.append("superseded")
    if record.source_id == "DOC-HB-003":
        flags.append("stale/conflicted: DOC-HB-009 says it replaces the March 2024 version")
    if record.source_id == "DOC-HB-006":
        flags.append("points readers to DOC-HB-006a")
    if "travel-and-subsistence-policy-v2.0" in record.source_path.name.lower() and "next review: april 2022" in blob:
        flags.append("past review: next review was April 2022")
    if "draft v0.8" in blob:
        flags.append("draft")
    if record.contains_synthetic_identifiers:
        flags.append("synthetic staff-directory fixture")
    if record.extraction_quality == "low":
        flags.append("low extraction quality")
    return unique(flags)


def classify_matches(records: list[SourceRecord]) -> None:
    for record in records:
        blob = f"{record.title}\n{record.extracted_markdown}".lower()
        for topic in TOPIC_DEFS:
            if topic["slug"] in record.topics or any(k in blob for k in topic["keywords"]):
                record.matched_topics.append(topic["slug"])
        record.matched_topics = unique(record.matched_topics)
        for entity in ENTITY_DEFS:
            if any(k in blob for k in entity["keywords"]):
                record.matched_entities.append(entity["slug"])
        record.matched_entities = unique(record.matched_entities)


def source_frontmatter(record: SourceRecord) -> dict[str, Any]:
    tags = unique(["source", "challenge-2", record.source_format, record.status] + record.topics + record.flags)
    return {
        "source_id": record.source_id,
        "title": record.title,
        "aliases": unique([record.source_id, record.title, record.source_path.stem]),
        "source_path": os.path.relpath(record.source_path, record.note_path.parent),
        "source_format": record.source_format,
        "document_type": record.document_type,
        "department": record.department,
        "owner": record.owner,
        "status": record.status,
        "version": record.version,
        "publication_date": record.publication_date,
        "last_updated": record.last_updated,
        "audience": record.audience,
        "topics": record.topics,
        "supersedes": record.supersedes,
        "related_sources": record.related_sources,
        "tags": tags,
        "data_origin": "synthetic_fixture",
        "extraction": {
            "method": record.extraction_method,
            "quality": record.extraction_quality,
            "warnings": record.extraction_warnings,
        },
        "sensitivity": {
            "contains_real_personal_data": False,
            "contains_synthetic_identifiers": record.contains_synthetic_identifiers,
            "classification": record.sensitivity_classification,
        },
    }


def truncate_sentences(text: str, max_chars: int = 650) -> str:
    clean = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    clean = re.sub(r"\s+", " ", re.sub(r"[#*_`|]+", " ", clean)).strip()
    if len(clean) <= max_chars:
        return clean
    return clean[:max_chars].rsplit(" ", 1)[0] + "..."


def clean_doc_link_label(label: str) -> str:
    label = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", label)
    return re.sub(r"\s+", " ", label).strip()


def doc_link(doc_id: str, record: SourceRecord, by_id: dict[str, SourceRecord], label: str | None = None) -> str:
    doc_id = doc_id.upper()
    target = by_id.get(doc_id)
    label = clean_doc_link_label(label or doc_id)
    if not target or target.source_id == record.source_id:
        return label
    return rel_link(record.note_path, target.note_path, label)


def normalise_markdown_doc_links(text: str, record: SourceRecord, by_id: dict[str, SourceRecord]) -> str:
    def repl(match: re.Match[str]) -> str:
        label, target = match.group(1), match.group(2)
        doc_match = DOC_ID_RE.search(f"{label} {target}".upper())
        if not doc_match:
            return match.group(0)
        return doc_link(doc_match.group(0), record, by_id, label)

    return MARKDOWN_LINK_RE.sub(repl, text)


def normalise_absolute_local_links(text: str) -> str:
    def repl(match: re.Match[str]) -> str:
        label, target = match.group(1), match.group(2)
        if target.startswith("/"):
            return clean_doc_link_label(label)
        return match.group(0)

    return MARKDOWN_LINK_RE.sub(repl, text)


def linkify(text: str, record: SourceRecord, by_id: dict[str, SourceRecord]) -> str:
    def repl(match: re.Match[str]) -> str:
        return doc_link(match.group(0), record, by_id)

    def guidance_repl(match: re.Match[str]) -> str:
        return doc_link(match.group(1), record, by_id)

    text = normalise_markdown_doc_links(text, record, by_id)
    text = re.sub(r"/guidance/(DOC-[A-Z]{2}-\d{3}[A-Za-z]?)\b", guidance_repl, text)
    text = normalise_absolute_local_links(text)

    parts: list[str] = []
    last = 0
    for match in MARKDOWN_LINK_RE.finditer(text):
        parts.append(DOC_ID_RE.sub(repl, text[last:match.start()]))
        parts.append(match.group(0))
        last = match.end()
    parts.append(DOC_ID_RE.sub(repl, text[last:]))
    return "".join(parts)


def metadata_table(record: SourceRecord) -> str:
    rows = [
        ("Source ID", record.source_id),
        ("Title", record.title),
        ("Raw source", rel_link(record.note_path, record.source_path, record.source_path.name)),
        ("Format", record.source_format),
        ("Document type", record.document_type),
        ("Department", record.department),
        ("Owner", record.owner),
        ("Status", record.status),
        ("Version", record.version),
        ("Publication date", record.publication_date),
        ("Last updated", record.last_updated),
        ("Audience", ", ".join(record.audience)),
        ("Topics", ", ".join(record.topics)),
        ("Data origin", SYNTHETIC_DATA_NOTICE),
        ("Contains real personal data", "false"),
        ("Contains synthetic identifiers", str(record.contains_synthetic_identifiers).lower()),
        ("SHA-256", record.sha256),
    ]
    return "\n".join(f"| {k} | {str(v or '').replace('|', '\\|')} |" for k, v in [("Field", "Value"), ("---", "---")] + rows)


def raw_metadata_section(record: SourceRecord) -> str:
    if not record.raw_metadata:
        return "_No embedded metadata was found._\n"
    rows = [("| Field | Value |"), ("| --- | --- |")]
    for key in sorted(record.raw_metadata):
        value = record.raw_metadata[key]
        if isinstance(value, (dict, list)):
            shown = json.dumps(value, ensure_ascii=False)
        else:
            shown = str(value)
        shown = shown.replace("\n", " ").replace("|", "\\|")
        rows.append(f"| {key} | {shown} |")
    return "\n".join(rows) + "\n"


def write_source_note(record: SourceRecord, by_id: dict[str, SourceRecord]) -> None:
    fm = source_frontmatter(record)
    bits: list[str] = [frontmatter(fm), f"# {record.title}\n"]
    bits.append("## Summary\n\n")
    bits.append(f"- Source: {rel_link(record.note_path, record.source_path, record.source_path.name)}\n")
    bits.append(f"- Extraction: `{record.extraction_method}` with `{record.extraction_quality}` quality.\n")
    if record.flags:
        bits.append(f"- Flags: {', '.join(record.flags)}.\n")
    bits.append(f"- Extract: {truncate_sentences(record.extracted_markdown)}\n\n")

    bits.append("## Metadata\n\n")
    bits.append(metadata_table(record) + "\n\n")

    if record.related_sources:
        bits.append("## Related Sources\n\n")
        for source_id in record.related_sources:
            target = by_id.get(source_id)
            if target:
                bits.append(f"- {rel_link(record.note_path, target.note_path, source_id)} - {target.title}\n")
            else:
                bits.append(f"- `{source_id}` - referenced but not found in this corpus\n")
        bits.append("\n")

    if record.matched_topics or record.matched_entities:
        bits.append("## Navigation\n\n")
        for topic_slug in record.matched_topics:
            topic = TOPICS_DIR / f"{topic_slug}.md"
            bits.append(f"- Topic: {rel_link(record.note_path, topic, next(t['title'] for t in TOPIC_DEFS if t['slug'] == topic_slug))}\n")
        for entity_slug in record.matched_entities:
            entity = ENTITIES_DIR / f"{entity_slug}.md"
            bits.append(f"- Entity: {rel_link(record.note_path, entity, next(e['title'] for e in ENTITY_DEFS if e['slug'] == entity_slug))}\n")
        bits.append("\n")

    if record.extraction_warnings:
        bits.append("## Extraction Warnings\n\n")
        for warning in record.extraction_warnings:
            bits.append(f"- {warning}\n")
        bits.append("\n")

    if record.extracted_tables:
        bits.append("## Extracted Tables\n\n")
        for index, table in enumerate(record.extracted_tables, start=1):
            bits.append(f"### Table {index}\n\n")
            bits.append(markdown_table(table["rows"]) + "\n")

    if record.extracted_links:
        bits.append("## Extracted Links\n\n")
        bits.append("| Label | URL | Note |\n| --- | --- | --- |\n")
        for link in record.extracted_links:
            label = link["label"].replace("|", "\\|")
            url = link["url"]
            note = link["note"].replace("|", "\\|")
            bits.append(f"| {label} | [{url}]({url}) | {note} |\n")
        bits.append("\n")

    bits.append("## Extracted Content\n\n")
    bits.append(linkify(record.extracted_markdown, record, by_id).strip() + "\n\n")

    bits.append("## Raw Metadata\n\n")
    bits.append(raw_metadata_section(record) + "\n")

    bits.append("## Related Notes\n\n")
    bits.append(f"- {rel_link(record.note_path, WIKI_DIR / 'index.md', 'Knowledge base index')}\n")
    bits.append(f"- {rel_link(record.note_path, WIKI_DIR / 'lint-report.md', 'Lint report')}\n")

    record.note_path.write_text("".join(bits), encoding="utf-8")


def source_table(records: list[SourceRecord], from_path: Path) -> str:
    lines = ["| Source | Status | Format | Department | Flags |", "| --- | --- | --- | --- | --- |"]
    for r in sorted(records, key=lambda x: x.source_id):
        flags = ", ".join(r.flags)
        lines.append(
            f"| {rel_link(from_path, r.note_path, r.source_id)} | {r.status or ''} | {r.source_format} | {r.department or ''} | {flags.replace('|', '\\|')} |"
        )
    return "\n".join(lines) + "\n"


def write_topic_notes(records: list[SourceRecord]) -> None:
    for topic in TOPIC_DEFS:
        note = TOPICS_DIR / f"{topic['slug']}.md"
        matched = [r for r in records if topic["slug"] in r.matched_topics]
        fm = {
            "title": topic["title"],
            "aliases": topic["aliases"],
            "note_type": "topic",
            "tags": ["topic", "challenge-2", topic["slug"]],
            "source_count": len(matched),
            "updated": TODAY,
        }
        bits = [frontmatter(fm), f"# {topic['title']}\n\n"]
        noun = "source document" if len(matched) == 1 else "source documents"
        bits.append(f"This topic page compiles {len(matched)} {noun} whose extracted content or metadata mentions this area.\n\n")
        if matched:
            bits.append("## Source Coverage\n\n")
            bits.append(source_table(matched, note) + "\n")
            bits.append("## Key Source Signals\n\n")
            for r in matched[:10]:
                bits.append(f"- {rel_link(note, r.note_path, r.source_id)}: {truncate_sentences(r.extracted_markdown, 220)}\n")
            bits.append("\n")
        else:
            bits.append("_No sources matched this topic in the current ingest._\n\n")
        bits.append("## Related Notes\n\n")
        for other in TOPIC_DEFS:
            if other["slug"] != topic["slug"]:
                shared = set(topic["keywords"]) & set(other["keywords"])
                if shared:
                    bits.append(f"- {rel_link(note, TOPICS_DIR / (other['slug'] + '.md'), other['title'])}\n")
        bits.append(f"- {rel_link(note, WIKI_DIR / 'index.md', 'Knowledge base index')}\n")
        note.write_text("".join(bits), encoding="utf-8")


def write_entity_notes(records: list[SourceRecord]) -> None:
    for entity in ENTITY_DEFS:
        note = ENTITIES_DIR / f"{entity['slug']}.md"
        matched = [r for r in records if entity["slug"] in r.matched_entities]
        fm = {
            "title": entity["title"],
            "aliases": entity["aliases"],
            "note_type": "entity",
            "tags": ["entity", "challenge-2", entity["slug"]],
            "source_count": len(matched),
            "updated": TODAY,
        }
        bits = [frontmatter(fm), f"# {entity['title']}\n\n"]
        bits.append(f"This entity appears in {len(matched)} source documents.\n\n")
        if matched:
            bits.append(source_table(matched, note) + "\n")
        bits.append("## Related Notes\n\n")
        for r in matched[:12]:
            bits.append(f"- {rel_link(note, r.note_path, r.source_id)}\n")
        bits.append(f"- {rel_link(note, WIKI_DIR / 'index.md', 'Knowledge base index')}\n")
        note.write_text("".join(bits), encoding="utf-8")


def write_maps(records: list[SourceRecord]) -> None:
    topic_lookup = {t["slug"]: t for t in TOPIC_DEFS}
    for map_def in MAP_DEFS:
        note = MAPS_DIR / f"{map_def['slug']}.md"
        topic_slugs = map_def["topics"]
        matched = [r for r in records if set(r.matched_topics) & set(topic_slugs)]
        fm = {
            "title": map_def["title"],
            "note_type": "map",
            "tags": ["map", "moc", "challenge-2", map_def["slug"]],
            "source_count": len(matched),
            "updated": TODAY,
        }
        bits = [frontmatter(fm), f"# {map_def['title']}\n\n"]
        bits.append("## Topics\n\n")
        for slug in topic_slugs:
            topic = topic_lookup[slug]
            bits.append(f"- {rel_link(note, TOPICS_DIR / (slug + '.md'), topic['title'])}\n")
        bits.append("\n## Sources\n\n")
        bits.append(source_table(matched, note) + "\n")
        flagged = [r for r in matched if r.flags]
        if flagged:
            bits.append("## Review Flags\n\n")
            for r in flagged:
                bits.append(f"- {rel_link(note, r.note_path, r.source_id)}: {', '.join(r.flags)}\n")
            bits.append("\n")
        bits.append("## Related Notes\n\n")
        bits.append(f"- {rel_link(note, WIKI_DIR / 'index.md', 'Knowledge base index')}\n")
        note.write_text("".join(bits), encoding="utf-8")


def write_data(records: list[SourceRecord]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    tables_readme = frontmatter({
        "title": "Generated Table Exports",
        "note_type": "data-readme",
        "tags": ["data", "challenge-2", "synthetic-fixtures"],
        "updated": TODAY,
    })
    tables_readme += "# Generated Table Exports\n\n"
    tables_readme += f"{SYNTHETIC_DATA_NOTICE} Names, email-like values, phone-like values, and job titles in these generated exports are synthetic identifiers retained for the Challenge 2 demo.\n\n"
    tables_readme += "Do still treat real secrets, local filesystem paths, and data from outside the Challenge 2 synthetic corpus as review issues.\n"
    (TABLES_DIR / "README.md").write_text(tables_readme, encoding="utf-8")
    register: list[dict[str, Any]] = []
    for r in records:
        register.append({
            "source_id": r.source_id,
            "title": r.title,
            "source_path": os.path.relpath(r.source_path, ROOT),
            "note_path": os.path.relpath(r.note_path, ROOT),
            "source_format": r.source_format,
            "document_type": r.document_type,
            "department": r.department,
            "status": r.status,
            "version": r.version,
            "publication_date": r.publication_date,
            "last_updated": r.last_updated,
            "audience": r.audience,
            "topics": r.topics,
            "matched_topics": r.matched_topics,
            "matched_entities": r.matched_entities,
            "supersedes": r.supersedes,
            "related_sources": r.related_sources,
            "flags": r.flags,
            "data_origin": "synthetic_fixture",
            "synthetic_data_notice": SYNTHETIC_DATA_NOTICE,
            "extraction": {
                "method": r.extraction_method,
                "quality": r.extraction_quality,
                "warnings": r.extraction_warnings,
                "table_count": len(r.extracted_tables),
                "links": r.extracted_links,
            },
            "sensitivity": {
                "contains_real_personal_data": False,
                "contains_synthetic_identifiers": r.contains_synthetic_identifiers,
                "classification": r.sensitivity_classification,
            },
            "technical_metadata": r.technical_metadata,
            "raw_metadata": r.raw_metadata,
        })
        if r.worksheets:
            workbook_json = {
                "source_id": r.source_id,
                "title": r.title,
                "source_path": os.path.relpath(r.source_path, ROOT),
                "data_origin": "synthetic_fixture",
                "synthetic_data_notice": SYNTHETIC_DATA_NOTICE,
                "contains_real_personal_data": False,
                "contains_synthetic_identifiers": r.contains_synthetic_identifiers,
                "worksheets": r.worksheets,
            }
            (TABLES_DIR / f"{slugify(r.source_id)}.json").write_text(json.dumps(workbook_json, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
            for ws in r.worksheets:
                csv_path = TABLES_DIR / f"{slugify(r.source_id)}-{slugify(ws['name'])}.csv"
                with csv_path.open("w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerows(ws["rows"])
    (DATA_DIR / "source-register.json").write_text(json.dumps(register, indent=2, ensure_ascii=False, default=str), encoding="utf-8")


def write_architecture_page(records: list[SourceRecord]) -> None:
    note = WIKI_DIR / "architecture.md"
    formats: dict[str, int] = {}
    for record in records:
        formats[record.source_format] = formats.get(record.source_format, 0) + 1
    flagged = [record for record in records if record.flags]
    xlsx_sources = [record for record in records if record.source_format == "xlsx"]
    fm = {
        "title": "Challenge 2 Architecture",
        "aliases": ["Architecture Overview", "How The Challenge 2 Wiki Works", "LLM Wiki Architecture"],
        "note_type": "architecture",
        "tags": ["architecture", "challenge-2", "llm-wiki", "start-here"],
        "source_count": len(records),
        "updated": TODAY,
    }
    bits = [frontmatter(fm), "# Challenge 2 Architecture\n\n"]
    bits.append("This page explains the knowledge-base architecture for someone seeing it for the first time. The short version: the original documents stay untouched, a repeatable builder reads them, and the generated wiki turns them into linked notes that are easy to browse, search, review, and demo.\n\n")

    bits.append("## What This Is\n\n")
    bits.append("Challenge 2 starts with a messy document collection: HTML, Markdown, text files, PDFs, Word documents, and spreadsheets. The architecture translates those documents into an Obsidian-friendly knowledge base without changing the source files. The result is a wiki with source notes, topic pages, entity pages, maps of content, a source register, and a lint report.\n\n")
    bits.append("The design follows the LLM Wiki pattern: raw sources are the source of truth, generated Markdown is the navigable knowledge layer, and explicit operating rules tell future agents how to maintain it.\n\n")

    bits.append("## System At A Glance\n\n")
    bits.append("```mermaid\n")
    bits.append("flowchart LR\n")
    bits.append("  Raw[\"Raw Challenge 2 documents\\nstructured_files and unstructured_files\"] --> Builder[\"build_wiki.py\\nrepeatable extractor and generator\"]\n")
    bits.append("  Builder --> Notes[\"Source notes\\none Markdown page per document\"]\n")
    bits.append("  Builder --> Data[\"Machine-readable data\\nsource register and table exports\"]\n")
    bits.append("  Builder --> Synthesis[\"Synthesis layer\\ntopics, entities, and maps\"]\n")
    bits.append("  Builder --> Lint[\"Lint report\\ncoverage, links, and known risks\"]\n")
    bits.append("  Notes --> Vault[\"Obsidian vault\\nopen challenge-2 folder\"]\n")
    bits.append("  Data --> Vault\n")
    bits.append("  Synthesis --> Vault\n")
    bits.append("  Lint --> Vault\n")
    bits.append("  Vault --> Users[\"Users\\nbrowse, search, follow links, demo answers\"]\n")
    bits.append("```\n\n")

    bits.append("## Ingest And Validation Flow\n\n")
    bits.append("```mermaid\n")
    bits.append("flowchart TD\n")
    bits.append("  Scan[\"Step 1: Scan corpus\\nfind all source files\"] --> Fingerprint[\"Step 2: Fingerprint\\nsize and SHA-256\"]\n")
    bits.append("  Fingerprint --> Extract[\"Step 3: Extract content\\nPandoc, pdftotext, ExifTool, openpyxl\"]\n")
    bits.append("  Extract --> Normalize[\"Step 4: Normalize metadata\\ntitle, status, dates, department, topics\"]\n")
    bits.append("  Normalize --> SourceNotes[\"Step 5: Write source notes\\nfrontmatter, raw metadata, provenance\"]\n")
    bits.append("  Normalize --> Tables[\"Step 6: Export tables\\nMarkdown, CSV, and JSON\"]\n")
    bits.append("  SourceNotes --> Link[\"Step 7: Cross-link\\nsource IDs, topics, entities, maps\"]\n")
    bits.append("  Tables --> Link\n")
    bits.append("  Link --> LintStep[\"Step 8: Lint\\ncoverage, broken links, known challenge flags\"]\n")
    bits.append("  LintStep --> Demo[\"Step 9: Demo-ready vault\\nindex, maps, source register, lint report\"]\n")
    bits.append("```\n\n")

    bits.append("## Knowledge Model\n\n")
    bits.append("```mermaid\n")
    bits.append("flowchart TB\n")
    bits.append("  Index[\"wiki/index.md\\nmain doorway\"] --> Architecture[\"wiki/architecture.md\\nplain-English architecture guide\"]\n")
    bits.append("  Index --> Maps[\"wiki/maps\\nmaps of content\"]\n")
    bits.append("  Maps --> Topics[\"wiki/topics\\npolicy and operational themes\"]\n")
    bits.append("  Maps --> Sources[\"wiki/sources\\none note per raw document\"]\n")
    bits.append("  Topics --> Sources\n")
    bits.append("  Entities[\"wiki/entities\\ndepartments, laws, forms, teams\"] --> Sources\n")
    bits.append("  Sources --> Raw[\"Raw files\\nimmutable source documents\"]\n")
    bits.append("  Sources --> Register[\"wiki/data/source-register.json\\nmachine-readable inventory\"]\n")
    bits.append("  Sources --> Tables[\"wiki/data/tables\\nspreadsheet exports\"]\n")
    bits.append("  Index --> Lint[\"wiki/lint-report.md\\nquality and risk signals\"]\n")
    bits.append("```\n\n")

    bits.append("## What The Builder Produces\n\n")
    bits.append(f"- `{len(records)}` source notes: one generated Markdown note for every Challenge 2 document.\n")
    bits.append(f"- `{len(TOPIC_DEFS)}` topic pages: reusable summaries for policy and operational themes.\n")
    bits.append(f"- `{len(ENTITY_DEFS)}` entity pages: departments, forms, laws, and named teams/programmes.\n")
    bits.append(f"- `{len(MAP_DEFS)}` maps of content: guided entry points for browsing related material.\n")
    bits.append(f"- `{len(xlsx_sources)}` workbook exports: each spreadsheet is preserved as Markdown tables, JSON, and CSV.\n")
    bits.append(f"- `{len(flagged)}` flagged sources: stale, draft, superseded, synthetic fixture identifiers, or past-review records.\n\n")

    bits.append("## Corpus Coverage\n\n")
    bits.append("| Format | Sources |\n| --- | ---: |\n")
    for source_format, count in sorted(formats.items()):
        bits.append(f"| `{source_format}` | {count} |\n")
    bits.append("\n")

    bits.append("## Demo Walkthrough\n\n")
    bits.append("1. Open the `challenge-2/` folder as the Obsidian vault.\n")
    bits.append(f"2. Start at {rel_link(note, WIKI_DIR / 'index.md', 'the knowledge base index')}.\n")
    bits.append(f"3. Use {rel_link(note, MAPS_DIR / 'housing-and-benefits.md', 'Housing And Benefits Map')} to show how policy documents connect.\n")
    bits.append(f"4. Use {rel_link(note, TOPICS_DIR / 'procurement-and-spending-controls.md', 'Procurement And Spending Controls')} to answer the IT hardware approval question.\n")
    bits.append(f"5. Use {rel_link(note, WIKI_DIR / 'lint-report.md', 'the lint report')} to show why metadata, provenance, and versioning matter.\n")
    bits.append(f"6. Use {rel_link(note, DATA_DIR / 'source-register.json', 'the source register')} to show the same knowledge base can also feed machine consumers.\n\n")

    bits.append("## Why The Architecture Matters\n\n")
    bits.append("- **Traceability:** every generated note links back to its raw source file.\n")
    bits.append("- **Repeatability:** the builder can regenerate the wiki from the source corpus.\n")
    bits.append("- **Findability:** maps, topics, entities, tags, and backlinks give multiple routes through the same material.\n")
    bits.append("- **Safety:** stale, superseded, draft, synthetic fixture identifiers, and sensitive classifications are highlighted instead of hidden.\n")
    bits.append("- **Synthetic fixtures:** all Challenge 2 raw and generated data is synthetic. Synthetic names and contact-like values are retained for demo fidelity; real secrets and local environment leaks remain review issues.\n")
    bits.append("- **Portability:** the output is plain Markdown and JSON, so it works in Obsidian, GitHub, VS Code, and simple scripts.\n\n")

    bits.append("## Glossary\n\n")
    bits.append("| Term | Meaning |\n| --- | --- |\n")
    glossary = [
        ("Architecture", "The way the raw files, extraction script, generated notes, data files, and Obsidian vault fit together."),
        ("Corpus", "The complete set of Challenge 2 source documents being processed."),
        ("Entity page", "A generated note about a department, law, form, team, or programme that appears across sources."),
        ("Extraction", "The process of reading content and metadata from source formats such as PDF, DOCX, HTML, Markdown, text, and XLSX."),
        ("Frontmatter", "YAML metadata at the top of a Markdown note, used by Obsidian and scripts for filtering and navigation."),
        ("Ingest", "One run of the builder that reads the corpus and regenerates the wiki layer."),
        ("Lint report", "A generated quality report covering source coverage, broken links, missing metadata, and known Challenge 2 risk signals."),
        ("LLM Wiki", "A maintained Markdown knowledge base that sits between raw sources and users, so knowledge compounds instead of being rediscovered from scratch."),
        ("Map of content", "A curated navigation page that groups related notes around a domain or workflow."),
        ("Obsidian vault", "A folder of Markdown files opened in Obsidian as a browsable knowledge base."),
        ("Provenance", "Evidence showing where a claim or extracted fact came from."),
        ("Raw source", "The original source document. In this architecture, raw sources are not edited."),
        ("Source note", "A generated Markdown note that represents one raw source file, including extracted text, metadata, links, and provenance."),
        ("Source register", "The machine-readable JSON inventory of every source file, extraction method, metadata fields, flags, and generated note path."),
        ("Synthetic fixture data", "Artificial data created for the Challenge 2 demo. It may look like staff or contact data, but it is not real personal data."),
        ("Topic page", "A generated synthesis note that groups sources around a recurring policy or operational theme."),
    ]
    for term, meaning in glossary:
        bits.append(f"| {term} | {meaning} |\n")
    bits.append("\n## Related Notes\n\n")
    bits.append(f"- {rel_link(note, WIKI_DIR / 'index.md', 'Knowledge base index')}\n")
    bits.append(f"- {rel_link(note, WIKI_DIR / 'lint-report.md', 'Lint report')}\n")
    bits.append(f"- {rel_link(note, ROOT / 'AGENTS.md', 'Operating rules')}\n")
    note.write_text("".join(bits), encoding="utf-8")


def write_index(records: list[SourceRecord]) -> None:
    note = WIKI_DIR / "index.md"
    by_format: dict[str, int] = {}
    for r in records:
        by_format[r.source_format] = by_format.get(r.source_format, 0) + 1
    fm = {
        "title": "Challenge 2 Knowledge Base Index",
        "aliases": ["Challenge 2 Wiki", "Dark Data Wiki"],
        "note_type": "index",
        "tags": ["index", "challenge-2", "llm-wiki"],
        "source_count": len(records),
        "updated": TODAY,
    }
    bits = [frontmatter(fm), "# Challenge 2 Knowledge Base Index\n\n"]
    bits.append("This is the navigation entrypoint for the Challenge 2 Obsidian knowledge base. Raw documents remain in `structured_files/` and `unstructured_files/`; generated knowledge lives under `wiki/`.\n\n")
    bits.append("## Start Here\n\n")
    bits.append(f"- {rel_link(note, WIKI_DIR / 'architecture.md', 'Architecture overview')}\n")
    bits.append(f"- {rel_link(note, WIKI_DIR / 'lint-report.md', 'Lint report')}\n")
    bits.append(f"- {rel_link(note, WIKI_DIR / 'log.md', 'Ingest log')}\n")
    bits.append(f"- {rel_link(note, DATA_DIR / 'source-register.json', 'Machine-readable source register')}\n\n")
    bits.append("## Maps Of Content\n\n")
    for m in MAP_DEFS:
        bits.append(f"- {rel_link(note, MAPS_DIR / (m['slug'] + '.md'), m['title'])}\n")
    bits.append("\n## Topic Pages\n\n")
    for topic in TOPIC_DEFS:
        count = len([r for r in records if topic["slug"] in r.matched_topics])
        bits.append(f"- {rel_link(note, TOPICS_DIR / (topic['slug'] + '.md'), topic['title'])} ({count} sources)\n")
    bits.append("\n## Entity Pages\n\n")
    for entity in ENTITY_DEFS:
        count = len([r for r in records if entity["slug"] in r.matched_entities])
        bits.append(f"- {rel_link(note, ENTITIES_DIR / (entity['slug'] + '.md'), entity['title'])} ({count} sources)\n")
    bits.append("\n## Source Corpus\n\n")
    bits.append(f"- Total sources: {len(records)}\n")
    for fmt, count in sorted(by_format.items()):
        bits.append(f"- `{fmt}`: {count}\n")
    bits.append("\n")
    bits.append(source_table(records, note) + "\n")
    bits.append("## Demo Questions\n\n")
    bits.append("- Which Council Tax Reduction guidance is current? Start with " + rel_link(note, TOPICS_DIR / "council-tax-reduction.md", "Council Tax Reduction") + ".\n")
    bits.append("- Can a self-employed person claim Housing Benefit? Start with " + rel_link(note, TOPICS_DIR / "small-business-and-self-employment.md", "Small Business And Self-employment") + " and " + rel_link(note, TOPICS_DIR / "housing-benefit.md", "Housing Benefit") + ".\n")
    bits.append("- Which staff policies are draft, stale, or past review? Start with " + rel_link(note, MAPS_DIR / "people-policies.md", "People Policies Map") + ".\n")
    bits.append("- What approvals are needed for IT hardware over GBP 5,000? Start with " + rel_link(note, TOPICS_DIR / "procurement-and-spending-controls.md", "Procurement And Spending Controls") + ".\n")
    bits.append("- Which documents mention Discretionary Housing Payments? Start with " + rel_link(note, TOPICS_DIR / "discretionary-housing-payments.md", "Discretionary Housing Payments") + ".\n")
    note.write_text("".join(bits), encoding="utf-8")


def write_log(records: list[SourceRecord]) -> None:
    note = WIKI_DIR / "log.md"
    if note.exists():
        existing = note.read_text(encoding="utf-8")
    else:
        existing = frontmatter({
            "title": "Challenge 2 Wiki Log",
            "note_type": "log",
            "tags": ["log", "challenge-2", "llm-wiki"],
        }) + "# Challenge 2 Wiki Log\n\n"
    flags = sum(1 for r in records if r.flags)
    entry = textwrap.dedent(f"""
    ## [{BUILD_TS}] ingest | Challenge 2 source corpus

    - Sources processed: {len(records)}
    - Structured sources: {len([r for r in records if 'structured_files' in r.source_path.parts])}
    - Unstructured sources: {len([r for r in records if 'unstructured_files' in r.source_path.parts])}
    - Source notes generated: {len(records)}
    - Flagged sources: {flags}
    - Source register: {rel_link(note, DATA_DIR / 'source-register.json', 'source-register.json')}
    - Lint report: {rel_link(note, WIKI_DIR / 'lint-report.md', 'lint-report.md')}

    """)
    note.write_text(existing.rstrip() + "\n\n" + entry.lstrip(), encoding="utf-8")


def md_files() -> list[Path]:
    return sorted(WIKI_DIR.rglob("*.md"))


def parse_frontmatter_exists(path: Path) -> bool:
    try:
        return path.read_text(encoding="utf-8").startswith("---\n")
    except UnicodeDecodeError:
        return False


def link_target_exists(source: Path, target: str) -> bool:
    target = target.strip()
    if re.match(r"^[a-z]+://", target) or target.startswith("#") or target.startswith("mailto:"):
        return True
    if target.startswith("/"):
        return False
    target = target.split("#", 1)[0]
    if not target:
        return True
    path = (source.parent / unquote(target)).resolve()
    return path.exists()


def lint(records: list[SourceRecord]) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    expected = sorted([p for d in [STRUCTURED_DIR, UNSTRUCTURED_DIR] for p in d.iterdir() if p.is_file()])
    if len(records) != len(expected):
        issues.append({"severity": "error", "message": f"Expected {len(expected)} source files but processed {len(records)}."})
    if len({r.note_path for r in records}) != len(records):
        issues.append({"severity": "error", "message": "Duplicate source note paths detected."})
    for r in records:
        if not r.note_path.exists():
            issues.append({"severity": "error", "source_id": r.source_id, "message": "Source note missing."})
        if not parse_frontmatter_exists(r.note_path):
            issues.append({"severity": "error", "source_id": r.source_id, "message": "Source note missing YAML frontmatter."})
        if not r.extraction_method or not r.source_path:
            issues.append({"severity": "error", "source_id": r.source_id, "message": "Missing extraction method or source path."})
        if not r.matched_topics:
            issues.append({"severity": "warning", "source_id": r.source_id, "message": "No topic match."})

    known_checks = {
        "DOC-HB-003": "stale/conflicted",
        "DOC-HB-006": "superseded",
        "UF-INFORMATION-SECURITY-POLICY-DRAFT-V0-8": "draft",
        "UF-TRAVEL-AND-SUBSISTENCE-POLICY-V2-0": "past review",
    }
    by_id = {r.source_id: r for r in records}
    for source_id, needle in known_checks.items():
        record = by_id.get(source_id)
        if not record or not any(needle in flag for flag in record.flags):
            issues.append({"severity": "error", "source_id": source_id, "message": f"Known challenge flag missing: {needle}."})

    for required in [
        "UF-SPENDING-CONTROLS-GUIDANCE",
        "UF-PROCUREMENT-THRESHOLDS-2024-25",
        "UF-STAFF-DIRECTORY-EXTRACT-Q4-2023",
        "UF-OVERPAYMENT-RECOVERY-PROCEDURES-V2-3",
    ]:
        if required not in by_id:
            issues.append({"severity": "error", "source_id": required, "message": "Required unstructured extraction missing."})

    for path in md_files():
        text = path.read_text(encoding="utf-8")
        for match in re.finditer(r"\[[^\]]+\]\(([^)]+)\)", text):
            target = match.group(1)
            if not link_target_exists(path, target):
                issues.append({"severity": "error", "file": os.path.relpath(path, ROOT), "message": f"Broken link: {target}"})

    report = {
        "generated_at": BUILD_TS,
        "source_count": len(records),
        "note_count": len(md_files()),
        "issue_count": len(issues),
        "issues": issues,
    }
    (DATA_DIR / "lint-report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return report


def write_lint_report(report: dict[str, Any], records: list[SourceRecord]) -> None:
    note = WIKI_DIR / "lint-report.md"
    fm = {
        "title": "Challenge 2 Wiki Lint Report",
        "note_type": "lint-report",
        "tags": ["lint", "challenge-2", "quality"],
        "generated_at": BUILD_TS,
        "issue_count": report["issue_count"],
    }
    bits = [frontmatter(fm), "# Challenge 2 Wiki Lint Report\n\n"]
    bits.append(f"- Generated at: `{BUILD_TS}`\n")
    bits.append(f"- Source count: {report['source_count']}\n")
    bits.append(f"- Note count: {report['note_count']}\n")
    bits.append(f"- Issue count: {report['issue_count']}\n\n")
    bits.append("## Known Challenge Checks\n\n")
    for r in records:
        if r.flags:
            bits.append(f"- {rel_link(note, r.note_path, r.source_id)}: {', '.join(r.flags)}\n")
    bits.append("\n## Issues\n\n")
    if report["issues"]:
        for issue in report["issues"]:
            bits.append(f"- `{issue['severity']}`: {issue.get('source_id') or issue.get('file') or ''} {issue['message']}\n")
    else:
        bits.append("_No lint issues found._\n")
    bits.append("\n## Related Notes\n\n")
    bits.append(f"- {rel_link(note, WIKI_DIR / 'index.md', 'Knowledge base index')}\n")
    bits.append(f"- {rel_link(note, DATA_DIR / 'lint-report.json', 'Machine-readable lint report')}\n")
    note.write_text("".join(bits), encoding="utf-8")


def clean_generated_dirs() -> None:
    for path in [SOURCES_DIR, TOPICS_DIR, ENTITIES_DIR, MAPS_DIR, DATA_DIR]:
        if path.exists():
            shutil.rmtree(path)
    for path in [SOURCES_DIR, TOPICS_DIR, ENTITIES_DIR, MAPS_DIR, TABLES_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def discover_sources() -> list[Path]:
    return sorted([p for d in [STRUCTURED_DIR, UNSTRUCTURED_DIR] for p in d.iterdir() if p.is_file()])


def build() -> dict[str, Any]:
    if not STRUCTURED_DIR.exists() or not UNSTRUCTURED_DIR.exists():
        raise SystemExit("Expected challenge-2/structured_files and challenge-2/unstructured_files to exist.")
    clean_generated_dirs()
    records = [extract_source(path) for path in discover_sources()]
    classify_matches(records)
    by_id = {r.source_id: r for r in records}
    for record in records:
        write_source_note(record, by_id)
    write_topic_notes(records)
    write_entity_notes(records)
    write_maps(records)
    write_data(records)
    write_architecture_page(records)
    write_index(records)
    write_log(records)
    (WIKI_DIR / "lint-report.md").write_text("# Pending lint report\n", encoding="utf-8")
    (DATA_DIR / "lint-report.json").write_text("{}\n", encoding="utf-8")
    report = lint(records)
    write_lint_report(report, records)
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Build the Challenge 2 Obsidian LLM wiki.")
    parser.add_argument("--strict", action="store_true", help="Return non-zero when lint errors are found.")
    args = parser.parse_args()
    report = build()
    errors = [i for i in report["issues"] if i["severity"] == "error"]
    print(f"Built Challenge 2 wiki: {report['source_count']} sources, {report['note_count']} notes, {report['issue_count']} lint issues.")
    if errors:
        print(f"Errors: {len(errors)}. See {WIKI_DIR / 'lint-report.md'}")
    if args.strict and errors:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
